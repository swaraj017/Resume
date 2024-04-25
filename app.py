# import os
# from PyPDF2 import PdfReader
# import spacy
#
# # Load English tokenizer, tagger, parser, NER, and word vectors
# nlp = spacy.load("en_core_web_sm")
#
# # Specify the full path to the folder containing resumes
# RESUME_FOLDER = "C:/Users/Swaraj Gaikwad/OneDrive/Desktop/Resume Short Lister/resumes"
#
# def extract_text_from_pdf(file_path):
#     text = ""
#     with open(file_path, "rb") as file:
#         pdf_reader = PdfReader(file)
#         for page_num in range(len(pdf_reader.pages)):
#             text += pdf_reader.pages[page_num].extract_text()
#     return text
#
# def analyze_resume(text):
#     doc = nlp(text)
#     # Example: score based on the number of tokens
#     return len(doc)
#
# def shortlist_resumes(resume_folder):
#     scores = {}
#     for filename in os.listdir(resume_folder):
#         if filename.endswith(".pdf"):
#             file_path = os.path.join(resume_folder, filename)
#             text = extract_text_from_pdf(file_path)
#             score = analyze_resume(text)
#             scores[filename] = score
#     # Sort resumes by score
#     sorted_resumes = sorted(scores.items(), key=lambda x: x[1], reverse=True)
#     return sorted_resumes
#
# if __name__ == "__main__":
#     # Assuming resumes are stored in the 'resumes' folder
#     resumes = shortlist_resumes(RESUME_FOLDER)
#     print("Shortlisted Resumes:")
#     for resume, score in resumes:
#         print(f"{resume}: Score - {score}")


# import os
# from PyPDF2 import PdfReader
# import spacy
# from sklearn.feature_extraction.text import TfidfVectorizer
# from sklearn.metrics.pairwise import cosine_similarity
#
# # Load English tokenizer, tagger, parser, NER, and word vectors
# nlp = spacy.load("en_core_web_sm")
#
# # Specify the full path to the folder containing resumes
# RESUME_FOLDER = "C:/Users/Swaraj Gaikwad/OneDrive/Desktop/Resume Short Lister/resumes"
#
#
# def extract_text_from_pdf(file_path):
#     text = ""
#     with open(file_path, "rb") as file:
#         pdf_reader = PdfReader(file)
#         for page_num in range(len(pdf_reader.pages)):
#             text += pdf_reader.pages[page_num].extract_text()
#     return text
#
#
# def preprocess_text(text):
#     # Tokenize and lemmatize the text
#     doc = nlp(text)
#     tokens = [token.lemma_ for token in doc if not token.is_stop and not token.is_punct]
#     return " ".join(tokens)
#
#
# def calculate_similarity(resume_text, job_description):
#     # Preprocess resume and job description text
#     preprocessed_resume = preprocess_text(resume_text)
#     preprocessed_job_description = preprocess_text(job_description)
#
#     # Create TF-IDF vectorizer
#     vectorizer = TfidfVectorizer()
#
#     # Fit and transform the text
#     tfidf_matrix = vectorizer.fit_transform([preprocessed_resume, preprocessed_job_description])
#
#     # Calculate cosine similarity
#     similarity_score = cosine_similarity(tfidf_matrix)[0, 1]
#
#     return similarity_score
#
#
# def shortlist_resumes(resume_folder, job_description):
#     scores = {}
#     for filename in os.listdir(resume_folder):
#         if filename.endswith(".pdf"):
#             file_path = os.path.join(resume_folder, filename)
#             resume_text = extract_text_from_pdf(file_path)
#             score = calculate_similarity(resume_text, job_description)
#             scores[filename] = score
#     # Sort resumes by score
#     sorted_resumes = sorted(scores.items(), key=lambda x: x[1], reverse=True)
#     return sorted_resumes
#
#
# if __name__ == "__main__":
#     # Assuming resumes are stored in the 'resumes' folder
#     job_description = "Your job description goes here"
#     resumes = shortlist_resumes(RESUME_FOLDER, job_description)
#     print("Shortlisted Resumes:")
#     for resume, score in resumes:
#         print(f"{resume}: Similarity Score - {score}")
# import os
# from PyPDF2 import PdfReader
# import spacy
#
# # Load English tokenizer, tagger, parser, NER, and word vectors
# nlp = spacy.load("en_core_web_sm")
#
# # Specify the full path to the folder containing resumes
# RESUME_FOLDER = "C:/Users/Swaraj Gaikwad/OneDrive/Desktop/Resume Short Lister/resumes"
#
#
# def extract_text_from_pdf(file_path):
#     text = ""
#     with open(file_path, "rb") as file:
#         pdf_reader = PdfReader(file)
#         for page_num in range(len(pdf_reader.pages)):
#             text += pdf_reader.pages[page_num].extract_text()
#     return text
#
#
# def preprocess_text(text):
#     # Tokenize and lemmatize the text
#     doc = nlp(text)
#     tokens = [token.lemma_ for token in doc if not token.is_stop and not token.is_punct]
#     return tokens
#
#
# def score_resume(tokens, job_keywords):
#     # Count the number of matching keywords
#     matching_keywords = set(tokens) & set(job_keywords)
#     return len(matching_keywords)
#
#
# def shortlist_resumes(resume_folder, job_description):
#     # Preprocess job description
#     job_keywords = preprocess_text(job_description)
#
#     scores = {}
#     for filename in os.listdir(resume_folder):
#         if filename.endswith(".pdf"):
#             file_path = os.path.join(resume_folder, filename)
#             resume_text = extract_text_from_pdf(file_path)
#             resume_tokens = preprocess_text(resume_text)
#             score = score_resume(resume_tokens, job_keywords)
#             scores[filename] = score
#     # Sort resumes by score
#     sorted_resumes = sorted(scores.items(), key=lambda x: x[1], reverse=True)
#     return sorted_resumes
#
#
# if __name__ == "__main__":
#     # Sample job description
#     job_description = """
#     We are looking for a Full Stack Web Developer proficient in Java, C++, JavaScript, Python, HTML5/CSS, and SQL.
#     Experience with web development frameworks such as React.js, Next.js, and Express.js is required.
#     Familiarity with version control systems like Git and databases including MongoDB, PostgreSQL, MySQL, and AWS is a plus.
#     """
#
#     # Assuming resumes are stored in the 'resumes' folder
#     resumes = shortlist_resumes(RESUME_FOLDER, job_description)
#     print("Shortlisted Resumes:")
#     for resume, score in resumes:
#         print(f"{resume}: Score - {score}")
# import os
# import re
# import spacy
# import json
# from PyPDF2 import PdfReader
# from collections import Counter
#
# # Load English tokenizer, tagger, parser, NER, and word vectors
# nlp = spacy.load("en_core_web_sm")
#
# # Specify the full path to the folder containing resumes
# RESUME_FOLDER = "C:/Users/Swaraj Gaikwad/OneDrive/Desktop/Resume Short Lister/resumes"
#
#
# def extract_skills(text):
#     # Define a regular expression pattern to match skills
#     skills_pattern = r'(Java|C\+\+|JavaScript|Python|HTML5/CSS|SQL|NodeJs|VScode|Git|Github|ReactJs|NextJs|ExpressJs|Tailwind CSS|MongoDB|PostgreSQL|MySQL|AWS)'
#     skills = re.findall(skills_pattern, text, re.IGNORECASE)
#     return [skill.lower() for skill in skills]
#
#
# def extract_text_from_pdf(file_path):
#     text = ""
#     with open(file_path, "rb") as file:
#         pdf_reader = PdfReader(file)
#         for page in pdf_reader.pages:
#             text += page.extract_text()
#     return text
#
#
# def extract_name_from_filename(filename):
#     # Extract the candidate's name from the filename
#     name = os.path.splitext(filename)[0]
#     name = name.replace("_", " ")  # Replace underscores with spaces
#     return name
#
#
# def score_resume(skills, job_keywords):
#     # Count the number of matching skills
#     skill_counter = Counter(skills)
#     score = sum(skill_counter[key] for key in job_keywords)
#     return score
#
#
# def shortlist_resumes(resume_folder, job_description):
#     # Extract skills from the job description
#     job_keywords = set(extract_skills(job_description))
#
#     shortlisted_resumes = []
#     for filename in os.listdir(resume_folder):
#         if filename.endswith(".pdf"):
#             file_path = os.path.join(resume_folder, filename)
#             text = extract_text_from_pdf(file_path)
#             skills = extract_skills(text)
#             score = score_resume(skills, job_keywords)
#             name = extract_name_from_filename(filename)
#             resume_data = {
#                 "name": name,
#                 "skills": skills,
#                 "score": score
#             }
#             shortlisted_resumes.append(resume_data)
#
#     # Sort resumes by score
#     sorted_resumes = sorted(shortlisted_resumes, key=lambda x: x["score"], reverse=True)
#     return sorted_resumes
#
#
# if __name__ == "__main__":
#     # Sample job description
#     job_description = """
#     We are looking for a Full Stack Web Developer proficient in Java, C++, JavaScript, Python, HTML5/CSS, and SQL.
#     Experience with web development frameworks such as React.js, Next.js, and Express.js is required.
#     Familiarity with version control systems like Git and databases including MongoDB, PostgreSQL, MySQL, and AWS is a plus.
#     """
#
#     # Assuming resumes are stored in the 'resumes' folder
#     shortlisted_resumes = shortlist_resumes(RESUME_FOLDER, job_description)
#     print("Shortlisted Resumes:")
#     for resume in shortlisted_resumes:
#         print(json.dumps(resume, indent=2))
import os
import re
import json
from PyPDF2 import PdfReader
import spacy
from openpyxl import Workbook
from collections import Counter
from datetime import datetime  # Add this line
from openpyxl import load_workbook
# Specify the full path to the folder containing resumes
RESUME_FOLDER = "C:/Users/Swaraj Gaikwad/OneDrive/Desktop/Resume Short Lister/resumes"

# Load English tokenizer, tagger, parser, NER, and word vectors
nlp = spacy.load("en_core_web_sm")


def extract_skills(text):
    # Define a regular expression pattern to match skills
    skills_pattern = r'(Java|C\+\+|JavaScript|Python|HTML5/CSS|SQL|NodeJs|VScode|Git|Github|ReactJs|NextJs|ExpressJs|Tailwind CSS|MongoDB|PostgreSQL|MySQL|AWS)'
    skills = re.findall(skills_pattern, text, re.IGNORECASE)
    return [skill.lower() for skill in skills]


def extract_text_from_pdf(file_path):
    text = ""
    with open(file_path, "rb") as file:
        pdf_reader = PdfReader(file)
        for page in pdf_reader.pages:
            text += page.extract_text()
    return text


def extract_name_from_filename(filename):
    # Extract the candidate's name from the filename
    name = os.path.splitext(filename)[0]
    name = name.replace("_", " ")  # Replace underscores with spaces
    return name


def score_resume(skills, job_keywords):
    # Count the number of matching skills
    skill_counter = Counter(skills)
    score = sum(skill_counter[key] for key in job_keywords)
    return score


def shortlist_resumes(resume_folder, job_description):
    # Extract skills from the job description
    job_keywords = set(extract_skills(job_description))

    shortlisted_resumes = []
    for filename in os.listdir(resume_folder):
        if filename.endswith(".pdf"):
            file_path = os.path.join(resume_folder, filename)
            text = extract_text_from_pdf(file_path)
            skills = extract_skills(text)
            score = score_resume(skills, job_keywords)
            name = extract_name_from_filename(filename)
            resume_data = {
                "name": name,
                "skills": skills,
                "score": score
            }
            shortlisted_resumes.append(resume_data)

    # Sort resumes by score
    sorted_resumes = sorted(shortlisted_resumes, key=lambda x: x["score"], reverse=True)
    return sorted_resumes


def create_excel_sheet(resumes):
    # Generate a unique filename based on current date and time
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"shortlisted_resumes_{current_datetime}.xlsx"

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add headers
    ws.append(["Name", "Skills", "Score"])

    # Add data
    for resume in resumes:
        ws.append([resume["name"], ", ".join(resume["skills"]), resume["score"]])

    # Save the workbook with the generated filename
    wb.save(excel_filename)
    print(f"Excel sheet created successfully: {excel_filename}")


if __name__ == "__main__":
    # Sample job description
    job_description = """
    We are looking for a Full Stack Web Developer proficient in Java, C++, JavaScript, Python, HTML5/CSS, and SQL.
    Experience with web development frameworks such as React.js, Next.js, and Express.js is required.
    Familiarity with version control systems like Git and databases including MongoDB, PostgreSQL, MySQL, and AWS is a plus.
    """

    # Assuming resumes are stored in the 'resumes' folder
    shortlisted_resumes = shortlist_resumes(RESUME_FOLDER, job_description)
    create_excel_sheet(shortlisted_resumes)


def excel_to_json(xlsx_file):
    data = []
    wb = load_workbook(filename=xlsx_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return json.dumps(data, indent=2)


def process_excel_files(excel_directory):
    excel_files = [filename for filename in os.listdir(excel_directory) if filename.endswith(".xlsx")]

    if excel_files:
        print("Excel files found:")
        print(excel_files)

        for excel_file in excel_files:
            excel_path = os.path.join(excel_directory, excel_file)
            json_data = excel_to_json(excel_path)
            print(f"\nJSON data for {excel_file}:")
            print(json_data)
    else:
        print("No Excel files found in the directory.")

if __name__ == "__main__":
    # excel_directory = "C:/Users/Swaraj Gaikwad/OneDrive/Desktop/Resume Short Lister/templates"
    # process_excel_files(excel_directory)
    print("Working")
    print("Shortlisted Resumes:")
    for resume in shortlisted_resumes:
            print(json.dumps(resume, indent=2))
